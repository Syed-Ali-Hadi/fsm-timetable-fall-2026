from http.server import BaseHTTPRequestHandler
from urllib.request import urlopen,Request
from io import BytesIO
from openpyxl import load_workbook
import json,re,unicodedata,datetime
SHEET_ID='1AnFQQhv9lu4grESE2ypbDG7E1QOPGgGCRiejem5ocPw'
DAYS={'Monday','Tuesday','Wednesday','Thursday','Friday'}
SLOTS=[(4,'08:30','09:50'),(13,'10:00','11:20'),(22,'11:30','12:50'),(31,'13:00','14:20'),(40,'14:25','15:45'),(49,'15:50','17:10')]
SECTION_RE=re.compile(r'^(?:BS)?(?:BBA|BA|AF|FT)\s*-?\s*\d{1,2}[A-Z](?:\d)?(?:\s*/\s*(?:(?:BS)?(?:BBA|BA|AF|FT))?\s*\d{0,2}[A-Z](?:\d)?)?$',re.I)
def fmt24(t):
 h,m=map(int,t.split(':'));return f"{h%12 or 12}:{m:02d} {'AM' if h<12 else 'PM'}"
def normalize_time(t):
 h,m=map(int,t.split(':'));h+=12 if 1<=h<=5 else 0;return f'{h:02d}:{m:02d}'
def ntext(v):
 s=unicodedata.normalize('NFKD',str(v or '')).encode('ascii','ignore').decode().lower().replace('&','and')
 return re.sub(r'[^a-z0-9]','',s)
def ncode(v):return re.sub(r'[^A-Z0-9/]','',str(v or '').upper())
def nsection(v):
 s=re.sub(r'[\s-]','',str(v or '').upper()).replace('BSFT','FT').replace('BSAF','AF')
 if s.startswith('BA') and not s.startswith('BBA'):s='BS'+s
 return s
def section_keys(v):
 s=nsection(v)
 if not s:return set()
 if '/' not in s:return {s}
 a,b=s.split('/',1);m=re.match(r'^(.*\d)([A-Z]\d?)$',a)
 return {a,m.group(1)+b} if re.fullmatch(r'[A-Z]\d?',b) and m else {a,b}
def is_section(v):return bool(SECTION_RE.fullmatch(re.sub(r'\s+','',str(v or ''))))
def course_plan(w):
 sheet=next((w[n] for n in w.sheetnames if n.strip().lower()=='course plan'),None);rows=[]
 if sheet:
  for r in range(1,sheet.max_row+1):
   code,title,section,teacher=sheet.cell(r,2).value,sheet.cell(r,3).value,sheet.cell(r,7).value,sheet.cell(r,8).value
   if code and title:rows.append({'code':ncode(code),'title':ntext(title),'sections':section_keys(section),'teacher':str(teacher or '').strip()})
 return rows
def choose_teacher(code,title,section,plan):
 code,title,sections=ncode(code),ntext(title),section_keys(section);scored=[]
 for row in plan:
  sec=bool(sections & row['sections']);samecode=bool(code and code==row['code']);sametitle=bool(title and (title==row['title'] or title in row['title'] or row['title'] in title))
  score=(6 if sec else 0)+(5 if samecode else 0)+(3 if sametitle else 0)
  if sec and (samecode or sametitle) and row['teacher']:scored.append((score,row['teacher']))
 if scored:
  best=max(x[0] for x in scored);names=sorted({x[1] for x in scored if x[0]==best})
  if len(names)==1:return names[0]
 if code:
  names=sorted({r['teacher'] for r in plan if r['code']==code and r['teacher']})
  if len(names)==1:return names[0]
 return 'TBA'
def slot_for_col(c):
 eligible=[x for x in SLOTS if x[0]<=c];return eligible[-1] if eligible else SLOTS[0]
def parse(blob):
 w=load_workbook(BytesIO(blob),data_only=True);s=w['Timetable'];plan=course_plan(w)
 merged_interior=set();merged_start={}
 for m in s.merged_cells.ranges:
  merged_start[(m.min_row,m.min_col)]=m.max_col
  for r in range(m.min_row,m.max_row+1):
   for c in range(m.min_col,m.max_col+1):
    if (r,c)!=(m.min_row,m.min_col):merged_interior.add((r,c))
 out=[];day='';typ='Class';recognized=0;skipped=[]
 for r in range(1,min(s.max_row,400)+1):
  first=str(s.cell(r,1).value or '').strip()
  if first in DAYS:day=first
  marker=str(s.cell(r,2).value or '').strip().lower()
  if marker.startswith('lab'):typ='Lab'
  elif marker.startswith('class'):typ='Class'
  room=str(s.cell(r,3).value or '').strip()
  if not day or not room or room.lower() in ('room','labs'):continue
  room=re.sub(r'\s+',' ',room).replace('A--','A-').replace('A- ','A-')
  anchors=[]
  for c in range(4,min(s.max_column,74)+1):
   if (r,c) in merged_interior:continue
   v=s.cell(r,c).value
   if v not in (None,''):anchors.append((c,str(v).strip()))
  for i,(sec_col,section) in enumerate(anchors):
   if not is_section(section):continue
   recognized+=1;j=i-1
   while j>=0 and is_section(anchors[j][1]):j-=1
   if j<0:skipped.append({'row':r,'col':sec_col,'reason':'no course','section':section});continue
   course_col,raw=anchors[j]
   if raw.lower() in ('cs','ms','jumma prayer') or raw.lower().startswith('seminar'):continue
   _,st,en=slot_for_col(course_col)
   times=re.findall(r'(\d{1,2}:\d{2})\s*[-–]\s*(\d{1,2}:\d{2})',raw)
   if times:
    st,en=map(normalize_time,times[-1]);raw=re.sub(r'\s*\(?\d{1,2}:\d{2}\s*[-–]\s*\d{1,2}:\d{2}\)?','',raw).strip()
   m=re.match(r'^([A-Z]{2}\s?\d{4})\s*',raw);code=m.group(1).replace(' ','') if m else '';title=raw[m.end():].strip() if m else raw
   teacher=choose_teacher(code,title,section,plan)
   semantic='|'.join(map(ntext,[day,st,en,room,code,title,section,typ]))
   out.append({'id':semantic,'day':day,'start':fmt24(st),'end':fmt24(en),'start24':st,'end24':en,'room':room,'type':typ,'code':code,'title':title,'section':re.sub(r'\s+','',section),'instructor':teacher})
 # Remove only exact semantic duplicates; preserve separate section occurrences.
 unique={}
 for x in out:unique.setdefault(x['id'],x)
 out=list(unique.values());matched=sum(x['instructor']!='TBA' for x in out)
 return {'entries':out,'meta':{'source':'Google Sheet: Timetable + Course Plan','live':True,'syncedAt':datetime.datetime.now(datetime.timezone.utc).isoformat(),'sourceSectionCells':recognized,'normalizedEntries':len(out),'teacherMatched':matched,'teacherTBA':len(out)-matched,'validationErrors':len(skipped),'validationDetails':skipped[:20]}}
class handler(BaseHTTPRequestHandler):
 def do_GET(self):
  try:
   urls=[f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx',f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx&gid=0']
   blob=None
   for u in urls:
    try:blob=urlopen(Request(u,headers={'User-Agent':'Mozilla/5.0'}),timeout=18).read();break
    except Exception:pass
   if not blob:raise RuntimeError('source unavailable')
   data=parse(blob)
   if data['meta']['validationErrors'] or data['meta']['normalizedEntries']<600:raise RuntimeError('validation failed')
   body=json.dumps(data).encode();self.send_response(200);self.send_header('Content-Type','application/json');self.send_header('Cache-Control','s-maxage=120, stale-while-revalidate=300');self.end_headers();self.wfile.write(body)
  except Exception:
   body=json.dumps({'error':'live source unavailable or failed validation'}).encode();self.send_response(503);self.send_header('Content-Type','application/json');self.end_headers();self.wfile.write(body)
